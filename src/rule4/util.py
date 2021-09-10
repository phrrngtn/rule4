from openpyxl.utils.cell import range_boundaries, coordinate_to_tuple
import pandas as pd

# probably best to always open the wobkbook with data_only=True.
# wb = openpyxl.load_workbook('foo.xlsx', data_only=True)

# otherwise can attempt something along these lines but it is unlikely to be without niggles
# https://stackoverflow.com/a/53787976/40387
# evaluate formula offline
def table_dict(wb):
    tables = {}
    for ws in wb.worksheets:
        for k,v in ws.tables.items():
            tables[k] = (ws.title,v)
    return tables


def named_ranges(wb):
    '''
    return dict of (sheet|None, range_name) -> ('TABLE'|'RANGE', name, destination)
    '''
    names = list([dn for dn in wb.defined_names.definedName])
    tables = table_dict(wb)
    ranges={}
    for n in filter(lambda x: x.type == 'RANGE' and  x.value.find('.xls')==-1, names):
        sn = None
        if n.localSheetId is not None:
             sn = wb.worksheets[n.localSheetId].title
        # TODO: extend this for named ranges that refer to structured references
        if n.value.find('[#All]')>0:
            #table_name = re.sub(r'\[.*\]$', '', n.value)
            table_name = n.value.replace('[#All]','')
            if not table_name in tables:
                print("Could not find table %s from %s" % (table_name, n.value))
            else:
                t = tables[table_name]
                ranges[(sn,n.name)] = ('TABLE', table_name, t)

        else:
            destination = list(n.destinations)[0]
            ranges[(sn,n.name)] =  ('RANGE', n.name, destination)
    return ranges

def cell_values(wb, sheet_name, rng):
    rb = range_boundaries(rng)
    ws = wb[sheet_name]
    #special-case scalars
    if rb[0]==rb[2] and rb[1]==rb[3]:
        return ws[rng].value
    else:
        return list(ws.iter_rows(min_col=rb[0],
                        min_row=rb[1],
                        max_col=rb[2],
                        max_row=rb[3],
                        values_only=True))

def extract_named_ranges(wb):
    results={}
    for k, v in named_ranges(wb).items():
        destination = v[2]
        values = cell_values(wb,destination[0], destination[1])
        if v[0] =='TABLE':
            results[k] =  pd.DataFrame(data=values[1:],columns=values[0]).dropna(how='all')
        else:
            results[k] =  values
    return results

def all_tables_as_dataframes(wb):
    '''return dictionary of table name to DataFrame'''
    tables = {}
    for sn in wb.sheetnames:
        sheetid = wb.sheetnames.index(sn)
        ws = wb.worksheets[sheetid]
        for tn in ws.tables:
            t = ws.tables[tn]
            rb = range_boundaries(t.ref)
            #print(sn, tn, rb)
            df = pd.DataFrame(data=ws.iter_rows(min_col=rb[0],
                                                # skip the header row.
                                                min_row=rb[1] + 1,
                                                max_col=rb[2],
                                                max_row=rb[3],
                                                values_only=True),  # just extract the value as a simple value rather than a Cell object
                            columns=[c.name for c in t.tableColumns],
                            ).dropna(how='all') # Get rid of blank rows.
            tables[tn] = df
    return tables

def all_table_references_as_dataframes(wb):
    '''
    Find all named ranges whose Target is a structured reference to an
    entire table (ListObject); accumulate them into a dictionary keyed by (sheet_name, named_range), values are dataframes
    constructed from the table contents. If the named range is Workbook-scoped, then sheet_name will be None
    '''
    tables =  all_tables_as_dataframes(wb)
    result = {}
    names = list([dn for dn in wb.defined_names.definedName])
    for n in names:
        if n.type == 'RANGE' and n.value.find('[#All]')>0:
            if n.value.find('.xls')>0:
                # This is a reference to a table in another workbook
                # so skip it.
                continue
            table_name = n.value.replace('[#All]', '')
            sn = None
            if n.localSheetId is not None:
                sn = wb.worksheets[n.localSheetId].title
            result[(sn, n.name)] = tables[table_name]
    return result

